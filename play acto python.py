from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import pathlib
from datetime import date
import tkinter as tk
import random
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
file = pathlib.Path('PRACEF.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet.title = "الوارد"
    sheet['A1'] = 'اسم الدواء'
    sheet['B1'] = 'سعر'
    sheet['C1']=' الوحده'
    sheet2 = file.create_sheet(title="الاصناف")
    sheet2['A1'] = 'user'
    sheet2['B1'] = 'password'
    file.save('PRACEF.xlsx')
#########################################################################################################################3

##########################################################################################################################3333




def cloose():
    root.destroy()

def returne():
    def cllose():
        root.destroy()
    root=Toplevel()
    Quant = StringVar()
    medcn = StringVar()
    unnt=StringVar()
    root.geometry('500x260')
    buy = Label(root, text='المرتجع', width=38, font=("Arial", 20), bg="white", fg="black")
    buy.grid(row=0, column=1, padx=5, pady=5, sticky='W')

    namee = Label(root, text='اسم الدواء', width=12, font=("Arial", 20), bg="white", fg="black")
    namee.grid(row=1, column=1, padx=280, pady=5, sticky=W)
    namee1 = Entry(root, text=medcn, width=16, font=("Arial", 20), bg="white", fg="black")
    namee1.grid(row=1, column=1, padx=20, pady=5, sticky=W)

    unit1 = Label(root, text='الوحده', width=12, font=("Arial", 20), bg="white", fg="black")
    unit1.grid(row=2, column=1, padx=280, pady=5, sticky=W)
    unit = ttk.Combobox(root, text=unnt, width=14, font=("Arial", 22))
    unit['values'] = ('اختار الكميه', 'علبه', 'شريط')
    unit.grid(row=2, column=1, padx=10, pady=5, sticky=W)
    unit.current(1)

    quantty = Label(root, text='الكميه', width=12, font=("Arial", 20), bg="white", fg="black")
    quantty.grid(row=3, column=1, padx=280, pady=5, sticky=W)
    quantty1 = Entry(root, text=Quant, width=16, font=("Arial", 20), bg="white", fg="black")
    quantty1.grid(row=3, column=1, padx=20, pady=5, sticky=W)

    upt = Button(root, text='اضافه', width=12, font=("Arial", 20), bg="white", fg="black")
    upt.grid(row=4, column=1, padx=280, pady=5, sticky=W)

    closs = Button(root, text='اغلاق', width=12, font=("Arial", 20), bg="white", fg="black",command=cllose)
    closs.grid(row=4, column=1, padx=60, pady=5, sticky=W)



    root.resizable(False, False)
    root.config(bg='cyan4')
######################### afirst
root=Tk()
root.geometry('1210x500')
root.title('OssAMCO')

def option_selected(event):
    global prce
    dor = name_e1.get().lower()
    wu = Unit.get()
    file = openpyxl.load_workbook('PRACEF.xlsx')
    sheet = file.active
    sheet.title = "الوارد"
    row=None
    for current_row in sheet.iter_rows(min_row=2,max_col=4,values_only=True):
        if  current_row[0]== str(dor):
            row= current_row
            if wu == 'شريط':
                print(f"Before prce.set: {prce.get()}")
                ##price1.delete(0, END)
                ##price1.insert(0, str(row[2]))
                print(f"after prce.set: {prce.get()}")
                prce.set(str(row[2]))
            elif  wu == 'علبه':
                print(f"Before prce.set: {prce.get()}")
                ##price1.delete(0, END)
                ##price1.insert(0, str(row[3]))
                print(f"after prce.set: {prce.get()}")
                prce.set(str(row[3]))
            else:
                messagebox.showerror("error", "تأكد من البيانات")  ###teratm
                price1.update_idletasks()
                break
            break
    else:
        messagebox.showerror("error", "تأكد من البيانات")###teratm
    file.close()
















Buy=Label(root,text='المبيعات',width=60,font=("Arial", 40), bg="white", fg="black",anchor='center')
Buy.grid(row=0,column=1,padx=10,pady=5,sticky='W')

medcin=StringVar()
name_e=Label(root,text='اسم الدواء',width=12,font=("Arial", 20), bg="white", fg="black")
name_e.grid(row=3,column=1,padx=760,pady=5,sticky=W)
name_e1=Entry(root,text=medcin,width=16,font=("Arial", 20), bg="white", fg="black")
name_e1.grid(row=3,column=1,padx=480,pady=5,sticky=W)

###Unt=StringVar()
Unit1=Label(root,text='الوحده',width=12,font=("Arial", 20), bg="white", fg="black")
Unit1.grid(row=3,column=1,padx=250,pady=5,sticky=W)
Unit=ttk.Combobox(root,text='Unt',width=12,font=("Arial", 22))
Unit['values']=('اختار الكميه','علبه','شريط')
Unit.grid(row=3,column=1,padx=10,pady=5,sticky=W)
Unit.current(0)
Unit.bind("<<ComboboxSelected>>", option_selected)

Quanty=StringVar()
Quantty=Label(root,text='الكميه',width=12,font=("Arial", 20), bg="white", fg="black")
Quantty.grid(row=4,column=1,padx=760,pady=5,sticky=W)
Quantty1=Entry(root,text=Quanty,width=16,font=("Arial", 20), bg="white", fg="black")
Quantty1.grid(row=4,column=1,padx=480,pady=5,sticky=W)


totl1=StringVar()
totl=Label(root,text='عدد الاصناف',width=12,font=("Arial", 20), bg="white", fg="black")
totl.grid(row=2,column=1,padx=250,pady=5,sticky=W)
total1=Entry(root,text=totl1,width=12,font=("Arial", 20), bg="white", fg="black")
total1.grid(row=2,column=1,padx=20,pady=5,sticky=W)

dat=StringVar()
dt=Label(root,text='التاريخ',width=12,font=("Arial", 20), bg="white", fg="black")
dt.grid(row=2,column=1,padx=760,pady=5,sticky=W)
dat1=Entry(root,text=dat,width=16,font=("Arial", 20), bg="white", fg="black")
dat1.grid(row=2,column=1,padx=480,pady=5,sticky=W)

today = date.today()
d2=today.strftime("%d/%m/%Y")
dat.set(d2)


prce = StringVar()
price=Label(root,text='السعر',width=12,font=("Arial", 20), bg="white", fg="black")
price.grid(row=4,column=1,padx=250,pady=5,sticky='w')
price1=Entry(root,textvariable=prce,width=12,font=("Arial", 20), bg="white", fg="black")
price1.grid(row=4,column=1,padx=20,pady=5,sticky=W)


discont=StringVar()
discount=Label(root,text='نسبه الخصم',width=12,font=("Arial", 20), bg="white", fg="black")
discount.grid(row=4,column=1,padx=250,pady=5,sticky='w')
discount1=Entry(root,text=discont,width=12,font=("Arial", 20), bg="white", fg="black")
discount1.grid(row=4,column=1,padx=20,pady=5,sticky=W)

priceyy=StringVar()
pricey=Label(root,text='السعر',width=12,font=("Arial", 20), bg="white", fg="black")
pricey.grid(row=5,column=1,padx=760,pady=5,sticky='w')
pricey1=Entry(root,text=priceyy,width=16,font=("Arial", 20), bg="white", fg="black")
pricey1.grid(row=5,column=1,padx=480,pady=5,sticky=W)

totalyy=StringVar()
totaly=Label(root,text='الاجمالى',width=12,font=("Arial", 20), bg="white", fg="black")
totaly.grid(row=6,column=1,padx=760,pady=5,sticky='w')
totaly1=Entry(root,text=totalyy,width=16,font=("Arial", 20), bg="white", fg="black")
totaly1.grid(row=6,column=1,padx=480,pady=5,sticky=W)


p=Button(root,text='اضافه',width=12,font=("Arial", 20), bg="white", fg="black")
p.grid(row=7,column=1,padx=760,pady=5,sticky=W)

cou=Button(root,text='احتساب',width=12,font=("Arial", 20), bg="white", fg="black")
cou.grid(row=7,column=1,padx=420,pady=5,sticky=W)

clos=Button(root,text='اغلاق',width=12,font=("Arial", 20), bg="white", fg="black",command=cloose)
clos.grid(row=7,column=1,padx=60,pady=5,sticky=W)

scher=Button(root,text='البحث',width=12,font=("Arial", 20), bg="white", fg="black")
scher.grid(row=2,column=1,padx=990,pady=5,sticky=W)

types=Button(root,text='الاصناف',width=12,font=("Arial", 20), bg="white", fg="black")
types.grid(row=5,column=1,padx=990,pady=5,sticky=W)

returned=Button(root,text='المرتجع',width=12,font=("Arial", 20), bg="white", fg="black",command=returne)
returned.grid(row=6,column=1,padx=990,pady=5,sticky=W)

import1=Button(root,text='الوارد',width=12,font=("Arial", 20), bg="white", fg="black")
import1.grid(row=3,column=1,padx=990,pady=5,sticky=W)

stoc=Button(root,text='المخزن',width=12,font=("Arial", 20), bg="white", fg="black")
stoc.grid(row=4,column=1,padx=990,pady=5,sticky=W)

root.resizable(False, False)
root.config(bg='cyan4')
root.mainloop()


